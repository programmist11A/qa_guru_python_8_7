import zipfile
import os.path
from PyPDF2 import PdfReader
import xlrd
from openpyxl import load_workbook



path_root = os.path.dirname(os.path.abspath(__file__))
path_tmp = os.path.join(path_root, 'tmp')
path_tmp_zip_file = os.path.join(path_root, 'tmp', 'arch.zip')

def test_txt():
    with zipfile.ZipFile(path_tmp_zip_file, 'r') as arch:
        with arch.open('text.txt') as text:
            assert 'hi' in text.read().decode('utf-8')


def test_pdf():
    with zipfile.ZipFile(path_tmp_zip_file, 'r') as arch:
        with arch.open('AByteofPythonRussian-2.02.pdf') as pdf_f:
            pdf_f = PdfReader(pdf_f)
            page = pdf_f.pages[55]
            text = page.extract_text()
            assert 'A Byte of Python' in text

            number_of_page = len(pdf_f.pages)
            print(number_of_page)
            assert number_of_page == 164

            x = arch.read('AByteofPythonRussian-2.02.pdf')
            pdf_file_size = len(x)
            print(pdf_file_size)
            expected_file_size = 794256
            assert pdf_file_size == expected_file_size


def test_xlsx():
    with zipfile.ZipFile(path_tmp_zip_file, 'r') as arch:
        with arch.open('Excel.xlsx') as xlsx_f:
            xlsx_file = load_workbook(xlsx_f)
            sheet = xlsx_file.active
            value = sheet.cell(row=7, column=2).value
            assert value == 'Фикси'


def test_xls():
    with zipfile.ZipFile(path_tmp_zip_file, 'r') as arch:
        with arch.open('sample-heavy-1.xls') as xls_f:
            xls_content = xls_f.read()
            xls_workbook = xlrd.open_workbook(file_contents=xls_content)
            sheet = xls_workbook.sheet_by_index(1)
            expected_value = 'run forest! RUN!!!'
            actual_value = sheet.cell_value(1, 1)
            assert actual_value == expected_value
